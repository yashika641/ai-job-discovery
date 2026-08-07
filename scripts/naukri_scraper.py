#!/usr/bin/env python
"""Standalone Playwright scraper for Naukri.com AI Engineer job listings.

Not wired into the daily pipeline (`pipeline.py`) yet -- this is a
standalone tool that opens naukri.com in a real browser, applies the
"Freshness" filter, scrapes the search results pages, and writes them
straight to an Excel file.

Freshness note: Naukri's own "Freshness" filter only offers fixed buckets
(Last 1 / 3 / 7 / 15 / 30 days) via the `jobAge` query param -- there's no
"at least 24h old" floor. "24 hours to 3 days max" is applied here as
`jobAge=3` ("Last 3 days"), the built-in bucket matching the 3-day ceiling;
each row's `posted_days_ago` / `posted_raw` columns show exactly how old
each listing actually is so you can eyeball/filter further in Excel.

Usage:
    python scripts/naukri_scraper.py
    python scripts/naukri_scraper.py --keyword "ai engineer" --location bengaluru
    python scripts/naukri_scraper.py --freshness 1 --max-pages 3 --show
    python scripts/naukri_scraper.py --output reports/naukri_ai_jobs.xlsx
"""

from __future__ import annotations

import argparse
import random
import re
import sys
import time
from dataclasses import dataclass, fields
from datetime import date
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from playwright.sync_api import Page, TimeoutError as PlaywrightTimeoutError, sync_playwright

REPO_ROOT = Path(__file__).resolve().parent.parent

JOB_CARD_SELECTORS = ["div.srp-jobtuple-wrapper", "article.jobTuple"]
BLOCK_PAGE_SIGNALS = ("access denied", "unusual traffic", "verify you are human", "captcha")

_POSTED_RE = re.compile(
    r"(?:(?P<num>\d+)\s*\+?\s*(?P<unit>day|days|hour|hours|week|weeks|month|months))"
    r"|(?P<today>today|just now|few hours ago)",
    re.IGNORECASE,
)


@dataclass
class JobRow:
    title: str = ""
    company: str = ""
    rating: str = ""
    experience: str = ""
    location: str = ""
    skills: str = ""
    description: str = ""
    posted_raw: str = ""
    posted_days_ago: str = ""
    job_url: str = ""


def parse_posted_days_ago(text: str) -> int | None:
    if not text:
        return None
    match = _POSTED_RE.search(text)
    if not match:
        return None
    if match.group("today"):
        return 0
    num = int(match.group("num"))
    unit = match.group("unit").lower()
    if unit.startswith("hour"):
        return 0
    if unit.startswith("day"):
        return num
    if unit.startswith("week"):
        return num * 7
    if unit.startswith("month"):
        return num * 30
    return None


def build_search_url(keyword: str, freshness: int, location: str | None) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", keyword.lower()).strip("-")
    url = f"https://www.naukri.com/{slug}-jobs?k={quote(keyword)}&jobAge={freshness}"
    if location:
        url += f"&l={quote(location)}"
    return url


def _first_matching_selector(page: Page, selectors: list[str], timeout: int) -> str | None:
    for selector in selectors:
        try:
            page.wait_for_selector(selector, timeout=timeout)
            return selector
        except PlaywrightTimeoutError:
            continue
    return None


def _dismiss_popups(page: Page) -> None:
    for selector in ('[title="Close"]', ".naukri-close-btn", "span.crossIcon", "button:has-text(\"Skip\")"):
        try:
            page.locator(selector).first.click(timeout=1500)
        except Exception:
            pass


def _check_not_blocked(page: Page) -> None:
    text = page.content().lower()
    if any(signal in text for signal in BLOCK_PAGE_SIGNALS):
        raise RuntimeError(
            "Naukri appears to be showing a bot-check / CAPTCHA page. Refusing to "
            "attempt a bypass -- open naukri.com in a normal browser, clear it "
            "manually, then re-run."
        )


def _text_or_empty(locator) -> str:
    try:
        return locator.first.inner_text(timeout=1000).strip()
    except Exception:
        return ""


def scrape_page(page: Page, card_selector: str) -> list[JobRow]:
    rows: list[JobRow] = []
    cards = page.locator(card_selector).all()
    for card in cards:
        try:
            title_el = card.locator("a.title").first
            title = _text_or_empty(title_el)
            job_url = title_el.get_attribute("href", timeout=1000) or ""
        except Exception:
            title, job_url = "", ""
        if not title:
            continue

        company = _text_or_empty(card.locator("a.comp-name, span.comp-name"))
        rating = _text_or_empty(card.locator(".rating span, span.main-2"))
        experience = _text_or_empty(card.locator(".exp-wrap, span.expwdth"))
        location = _text_or_empty(card.locator(".loc-wrap, span.locWdth"))
        description = _text_or_empty(card.locator("span.job-desc, .job-desc"))
        posted_raw = _text_or_empty(card.locator("span.job-post-day"))

        skill_els = card.locator("ul.tags-gt li, ul.tags li")
        try:
            skills = ", ".join(t.strip() for t in skill_els.all_inner_texts() if t.strip())
        except Exception:
            skills = ""

        days_ago = parse_posted_days_ago(posted_raw)
        rows.append(
            JobRow(
                title=title,
                company=company,
                rating=rating,
                experience=experience,
                location=location,
                skills=skills,
                description=description,
                posted_raw=posted_raw,
                posted_days_ago=str(days_ago) if days_ago is not None else "",
                job_url=job_url,
            )
        )
    return rows


def go_to_next_page(page: Page, card_selector: str) -> bool:
    next_link = page.locator("a.styles_btn-secondary__2AsIP, a", has_text="Next").last
    try:
        if next_link.count() == 0 or not next_link.is_enabled():
            return False
        first_title_before = _text_or_empty(page.locator(f"{card_selector} a.title").first)
        next_link.click(timeout=5000)
    except Exception:
        return False

    try:
        page.wait_for_function(
            """([selector, prevTitle]) => {
                const el = document.querySelector(selector + ' a.title');
                return el && el.innerText.trim() !== prevTitle;
            }""",
            arg=[card_selector, first_title_before],
            timeout=15000,
        )
    except PlaywrightTimeoutError:
        return False
    return True


def scrape_naukri(
    keyword: str,
    freshness: int,
    location: str | None,
    max_pages: int,
    headless: bool,
) -> list[JobRow]:
    url = build_search_url(keyword, freshness, location)
    print(f"Opening: {url}")

    all_rows: list[JobRow] = []
    seen_urls: set[str] = set()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1440, "height": 900},
        )
        page = context.new_page()
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=30000)
            _dismiss_popups(page)
            _check_not_blocked(page)

            card_selector = _first_matching_selector(page, JOB_CARD_SELECTORS, timeout=20000)
            if card_selector is None:
                print("No job cards found -- Naukri's markup may have changed, or there are 0 results.")
                return []

            for page_num in range(1, max_pages + 1):
                page.wait_for_timeout(1500)
                rows = scrape_page(page, card_selector)
                new_rows = [r for r in rows if r.job_url not in seen_urls]
                for r in new_rows:
                    seen_urls.add(r.job_url)
                all_rows.extend(new_rows)
                print(f"  page {page_num}: {len(rows)} cards, {len(new_rows)} new (total so far: {len(all_rows)})")

                if page_num == max_pages:
                    break
                if not go_to_next_page(page, card_selector):
                    print("  no more pages.")
                    break
                time.sleep(random.uniform(2, 4))
        finally:
            context.close()
            browser.close()

    return all_rows


def write_excel(rows: list[JobRow], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Naukri AI Engineer Jobs"

    headers = [f.name for f in fields(JobRow)]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([getattr(row, h) for h in headers])

    url_col = headers.index("job_url") + 1
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=url_col)
        if cell.value:
            cell.hyperlink = cell.value
            cell.font = Font(color="0563C1", underline="single")

    widths = {"title": 40, "company": 25, "description": 60, "skills": 35, "job_url": 45}
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 16)

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--keyword", default="ai engineer")
    parser.add_argument("--location", default=None)
    parser.add_argument(
        "--freshness", type=int, default=3, choices=[1, 3, 7, 15, 30],
        help="Naukri 'Last N days' bucket to filter by (default: 3, i.e. within 3 days)",
    )
    parser.add_argument("--max-pages", type=int, default=5)
    parser.add_argument("--show", action="store_true", help="Run with a visible browser window instead of headless")
    parser.add_argument(
        "--output", default=None,
        help="Excel output path (default: reports/naukri_ai_engineer_jobs_<today>.xlsx)",
    )
    args = parser.parse_args()

    output_path = Path(args.output) if args.output else REPO_ROOT / "reports" / f"naukri_ai_engineer_jobs_{date.today().isoformat()}.xlsx"

    try:
        rows = scrape_naukri(
            keyword=args.keyword,
            freshness=args.freshness,
            location=args.location,
            max_pages=args.max_pages,
            headless=not args.show,
        )
    except RuntimeError as exc:
        print(f"Stopped: {exc}")
        sys.exit(1)

    if not rows:
        print("No jobs scraped -- nothing written.")
        return

    write_excel(rows, output_path)
    print(f"\nWrote {len(rows)} jobs to {output_path}")


if __name__ == "__main__":
    main()
